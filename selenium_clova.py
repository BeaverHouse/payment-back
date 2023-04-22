from selenium import webdriver
from selenium.webdriver import ChromeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
import os
import glob
import openpyxl
import calendar
import shutil
from pathlib import Path
from parse_text import parse_text

"""
파일 처리 전체 프로세스
"""
def file_ocr(name, year, month):

    # 옵션 생성
    options = ChromeOptions()
    # 창 숨기는 옵션 추가
    options.add_argument("headless")

    driver = webdriver.Chrome('./chromedriver', options=options)

    # CLOVA 체험 사이트 접속
    driver.get('https://clova.ai/ocr/?lang=ko')
    driver.implicitly_wait(5)

    # 로컬 스토리지 초기화하고 업로드 버튼 클릭
    driver.execute_script("window.localStorage.clear();")
    driver.find_element(By.CLASS_NAME, "btn_upload").click()
    driver.implicitly_wait(5)

    # 동의 포맷 작성 (대충 입력하면 됨)
    driver.find_element(By.CLASS_NAME, "label_check").click()
    driver.find_element(By.ID, "name_terms").send_keys("a")
    driver.find_element(By.ID, "company_terms").send_keys("a") 
    driver.find_element(By.ID, "email_terms").send_keys("example@test.com") 

    # 스크롤 내려서 확인 버튼 클릭
    # popup 클래스가 2개 : 하나는 동의 포맷 팝업, 하나는 이미지 업로드 팝업
    body = driver.find_element(By.TAG_NAME, "body")
    body.send_keys(Keys.ENTER)
    for _ in range(5):
        body.send_keys(Keys.DOWN)
        
    time.sleep(2)

    popups = driver.find_elements(By.CLASS_NAME, "popup")
    popups[0].find_element(By.CLASS_NAME, "btn").click()

    time.sleep(2)

    # 일단은 X 버튼 눌러서 원래 화면으로 (반복실행을 위해)
    popups[1].find_element(By.CLASS_NAME, "btn_close").click()

    file_list = glob.glob("screenshots/{}_{}_{}/*".format(name, year, month))
    file_list_img = [file for file in file_list if (file.endswith(".jpg") or file.endswith(".png"))]

    arr = []
    for f in file_list_img:
        # 파일 업로드하여 제출
        driver.find_element(By.CLASS_NAME, "btn_upload").click()
        driver.implicitly_wait(5)
        driver.find_element(By.ID,"input_file").send_keys(os.path.abspath(f))

        time.sleep(2)
        popups[1].find_element(By.CLASS_NAME, "btn").click()

        # 업로드 대기
        time.sleep(10)

        # 텍스트 읽기
        result = body.find_element(By.CLASS_NAME, "data_result_area").text
        arr += parse_text(result)

    time.sleep(5)
    driver.quit()

    # sort
    arr = sorted(arr, key=lambda x: x["day"])
    arr = sorted(arr, key=lambda x: x["month"])
    
    
    # 엑셀 작성
    wb = openpyxl.load_workbook('format.xlsx')
    sh = wb["사용내역"]
    sh.cell(row=2, column=2).value = '{}월 법인카드 사용내역서'.format(month)
    sh.cell(row=3, column=2).value = name
    sh.cell(row=3, column=5).value = '기     간   :    {} / {} / 01  ~  {} / {} / {}'.format(
        year, month, year, month, calendar.monthrange(int(year), int(month))[1]
    )

    row = 5
    for a in arr:
        if a["month"] != month:
            continue
        sh.cell(row=row, column=3).value = "{}월 {}일".format(a["month"], a["day"])
        sh.cell(row=row, column=4).value = a["shop_name"]
        sh.cell(row=row, column=5).value = a["price"]
        sh.cell(row=row, column=7).value = "복리후생비"
        sh.cell(row=row, column=8).value = "CloudNetworks"
        sh.merge_cells(start_row=row, start_column=9, end_row=row, end_column=10)
        # sh.cell(row=row, column=9).value = "법인카드"
        row += 1

    Path("result").mkdir(parents=True, exist_ok=True)
    wb.save('result/{} {}년 {}월 법인카드 내역.xlsx'.format(name, year, month))


    # 처리된 이미지는 삭제
    shutil.rmtree("screenshots/{}_{}_{}".format(name, year, month), )


    # 남은 목록이 있으면 오래된 것부터 실행
    paths = sorted(Path("screenshots").iterdir(), key=os.path.getmtime)
    if len(paths) > 0:
        name, year, month = paths[0].name.split("_")
        with open("process.txt", "w", encoding="utf8") as f:
            f.write(paths[0].name)
        file_ocr(name, year, month)
    else:
        with open("process.txt", "w", encoding="utf8") as f:
            f.write("Idle")